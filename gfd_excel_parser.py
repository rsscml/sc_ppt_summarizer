"""
GFD Excel Parser Module
========================
Robust, schema-aware parser for the `Dashboard_Update` worksheet.

Designed for real-world Excel files that may have:
  - Multi-row / multi-level headers (category row + column name row)
  - Headers that don't start at row 1 (title rows, logos, blank rows above)
  - Merged cells in headers (horizontal and vertical spans)
  - Inconsistent casing, extra whitespace, special characters in headers
  - Subtotal / separator rows mixed with data
  - Trailing empty columns beyond the real data boundary
  - Columns in any order

Design principles:
  - Warn, never crash — missing/malformed columns emit warnings
  - No hard-coded column positions — fuzzy header matching only
  - Schema-flexible — tolerates extra or missing columns
  - Read-only — never modifies the source file
"""

import re
from pathlib import Path
from datetime import datetime, date
from typing import Any, Optional
from difflib import SequenceMatcher

import openpyxl


# ─── Column Mapping ──────────────────────────────────────────────────
# Maps semantic field names → list of possible header variations (lowered).
# These are matched using containment + fuzzy similarity, not just equality.

COLUMN_MAP: dict[str, list[str]] = {
    "basis":                    ["basis"],
    "link":                     ["link"],
    "customer_affected":        ["customer affected"],
    "ops_capacity_risk":        ["ops capacity risk", "operational capacity risk",
                                 "ops risk"],
    "strategic_capacity_risk":  ["strategic capacity risk", "strategic risk"],
    "region":                   ["region"],
    "plant_location":           ["plant / location", "plant/location", "plant",
                                 "location", "site"],
    "product_family_code":      ["product family code", "pf code", "product code",
                                 "pg code", "product group code"],
    "product_family_desc":      ["product family description", "product family desc",
                                 "pf description", "product family",
                                 "product group", "pg description", "pg"],
    "critical_component":       ["critical component", "component"],
    "constraint_task_force":    ["constraint / task force", "constraint/task force",
                                 "constraint", "task force"],
    "root_cause":               ["root cause", "cause"],
    "vendor_code":              ["vendor code", "vendor"],
    "supplier_number":          ["supplier number", "supplier no", "supplier id",
                                 "supplier #"],
    "supplier_text":            ["supplier text", "supplier name", "supplier"],
    "supplier_type":            ["supplier (internal / external)", "supplier type",
                                 "internal / external", "internal/external",
                                 "int / ext", "int/ext"],
    "supplier_region":          ["supplier region (local / overseas / regional)",
                                 "supplier region", "local / overseas / regional",
                                 "local/overseas", "supplier geo"],
    "coverage_without_mitigation": ["current coverage w/o risk mitigation",
                                    "coverage w/o risk mitigation",
                                    "coverage without mitigation",
                                    "coverage w/o mitigation",
                                    "coverage w/o", "coverage without"],
    "coverage_with_mitigation":    ["coverage w/ risk mitigation",
                                    "coverage with risk mitigation",
                                    "coverage with mitigation",
                                    "coverage w/ mitigation",
                                    "coverage w/", "coverage with"],
    "fulfillment_current_q":       ["% of customer order fulfillment - current q",
                                    "fulfillment - current q", "fulfillment current q",
                                    "customer order fulfillment - current q",
                                    "fulfillment current", "order fulfillment current"],
    "fulfillment_q_plus_1":        ["% of customer order fulfillment - q+1",
                                    "fulfillment - q+1", "fulfillment q+1",
                                    "customer order fulfillment - q+1",
                                    "fulfillment q1", "order fulfillment q+1"],
    "fulfillment_q_plus_2":        ["% of customer order fulfillment - q+2",
                                    "fulfillment - q+2", "fulfillment q+2",
                                    "customer order fulfillment - q+2",
                                    "fulfillment q2", "order fulfillment q+2"],
    "recovery_week":               ["recovery week (date - out of backlog)",
                                    "recovery week", "recovery date",
                                    "recovery", "out of backlog"],
    "last_update":              ["last update", "last updated", "updated"],
    "allocation_mode":          ["allocation mode [y/n]", "allocation mode",
                                 "allocation mode y/n"],
    "customer_informed":        ["customer informed [y/n]", "customer informed",
                                 "customer informed y/n", "cust informed"],
    "action_comment":           ["action / comment", "action/comment",
                                 "action comment", "comment", "action",
                                 "actions / comments", "actions/comments"],
    "task_force_leader":        ["task force leader", "tf leader", "tf lead"],
    "location_contact":         ["location contact", "site contact",
                                 "plant contact"],
    "special_freight_cost":     ["special freight cost [€]", "special freight cost",
                                 "freight cost", "special freight cost [eur]",
                                 "freight cost €", "freight cost eur"],
    "freight_cost_checked":     ["special freight cost coverage checked? [y/n]",
                                 "freight cost coverage checked",
                                 "cost coverage checked",
                                 "freight coverage checked"],
    "special_freight_remarks":  ["special freight"],
    "allocation":               ["allocation"],
    "allocation_responsible":   ["allocation responsible",
                                 "allocation resp"],
}

# These fields are disambiguation-sensitive: a short header like "allocation"
# should NOT match "allocation responsible", so we check exact-first for them.
_EXACT_FIRST_FIELDS = {"allocation", "supplier_text"}

# Known non-data row markers (subtotals, separators, etc.)
_SKIP_ROW_MARKERS = re.compile(
    r'^\s*(total|subtotal|sum|grand total|section\s+\d|---|===)\s*$',
    re.IGNORECASE
)


# ─── Calendar Week (CW) Parsing ─────────────────────────────────────

def parse_cw(value: Any, default_year: Optional[int] = None) -> Optional[tuple[int, int]]:
    """
    Parse a calendar-week string into (year, week_number).

    Accepted formats:
      "CW18"          → (default_year, 18)
      "CW18/2026"     → (2026, 18)
      "CW18/26"       → (2026, 18)
      "CW 22"         → (default_year, 22)
      "18"            → (default_year, 18)
      "W18"           → (default_year, 18)
      "KW18"          → (default_year, 18)   (German: Kalenderwoche)
      "CW18-2026"     → (2026, 18)

    Returns None if unparseable.
    """
    if value is None:
        return None
    if default_year is None:
        default_year = datetime.now().year

    s = str(value).strip().upper()
    if not s:
        return None

    # Strip common prefixes and whitespace
    m = re.match(r'^(?:CW|KW|W)?\s*(\d{1,2})\s*(?:[/\-\.]\s*(\d{2,4}))?\s*$', s)
    if not m:
        return None

    week = int(m.group(1))
    if week < 1 or week > 53:
        return None

    if m.group(2):
        year = int(m.group(2))
        if year < 100:
            year += 2000
    else:
        year = default_year

    return (year, week)


def cw_to_absolute(year: int, week: int) -> int:
    """Convert (year, week) to a monotonically increasing integer for comparison."""
    return year * 53 + week


def absolute_to_cw(absolute: int) -> tuple[int, int]:
    """Convert absolute week number back to (year, week)."""
    year = absolute // 53
    week = absolute % 53
    if week == 0:
        year -= 1
        week = 53
    return (year, week)


def get_current_cw() -> tuple[int, int]:
    """Get the current ISO calendar week as (year, week)."""
    today = datetime.now()
    iso = today.isocalendar()
    return (iso[0], iso[1])


# ─── Value Normalisation Helpers ─────────────────────────────────────

def _norm_str(val: Any) -> Optional[str]:
    if val is None:
        return None
    s = str(val).strip()
    return s if s else None


def _norm_pct(val: Any) -> Optional[float]:
    if val is None:
        return None
    if isinstance(val, (int, float)):
        return float(val) if val <= 1.0 else float(val) / 100.0
    s = str(val).strip().rstrip("%").strip()
    if not s:
        return None
    try:
        v = float(s.replace(",", "."))
        return v / 100.0 if v > 1.0 else v
    except ValueError:
        return None


def _norm_bool(val: Any) -> Optional[bool]:
    if val is None:
        return None
    s = str(val).strip().lower()
    if s in ("y", "yes", "true", "1", "x", "ja"):
        return True
    if s in ("n", "no", "false", "0", "", "nein"):
        return False
    return None


def _norm_date(val: Any) -> Optional[str]:
    if val is None:
        return None
    if isinstance(val, (datetime, date)):
        return val.isoformat()[:10]
    s = str(val).strip()
    if not s:
        return None
    # Try common date formats
    for fmt in ("%Y-%m-%d", "%d.%m.%Y", "%d/%m/%Y", "%m/%d/%Y", "%d-%m-%Y"):
        try:
            return datetime.strptime(s, fmt).date().isoformat()
        except ValueError:
            continue
    return s  # preserve as-is if no format matched


def _norm_float(val: Any) -> Optional[float]:
    if val is None:
        return None
    if isinstance(val, (int, float)):
        return float(val)
    s = str(val).strip()
    # Remove currency symbols and thousands separators
    s = re.sub(r'[€$£¥\s]', '', s)
    # Handle European number format: 1.234,56 → 1234.56
    if ',' in s and '.' in s:
        if s.index(',') > s.index('.'):
            s = s.replace('.', '').replace(',', '.')
        else:
            s = s.replace(',', '')
    elif ',' in s:
        s = s.replace(',', '.')
    if not s:
        return None
    try:
        return float(s)
    except ValueError:
        return None


_NORMALISERS = {
    "fulfillment_current_q":       _norm_pct,
    "fulfillment_q_plus_1":        _norm_pct,
    "fulfillment_q_plus_2":        _norm_pct,
    "last_update":                 _norm_date,
    "allocation_mode":             _norm_bool,
    "customer_informed":           _norm_bool,
    "freight_cost_checked":        _norm_bool,
    "special_freight_cost":        _norm_float,
}


# ─── Text Normalisation for Header Matching ──────────────────────────

def _normalise_header_text(text: str) -> str:
    """
    Aggressively normalise header text for matching.
    Strips whitespace, newlines, collapses runs, lowercases.
    """
    if not text:
        return ""
    s = text.lower()
    s = re.sub(r'[\n\r\t]+', ' ', s)        # newlines → space
    s = re.sub(r'[_]+', ' ', s)              # underscores → space
    s = re.sub(r'\s+', ' ', s)               # collapse whitespace
    s = s.strip()
    return s


def _fuzzy_score(candidate: str, target: str) -> float:
    """Return similarity ratio between candidate and target (0.0 – 1.0)."""
    if not candidate or not target:
        return 0.0
    # Exact containment is a strong signal
    if target in candidate or candidate in target:
        return 0.95
    return SequenceMatcher(None, candidate, target).ratio()


# ─── Multi-Row Header Detection & Flattening ─────────────────────────

def _get_merged_cell_map(ws) -> dict[tuple[int, int], Any]:
    """
    Build a lookup: (row, col) → master cell value for ALL merged ranges.
    Both row and col are 1-indexed (openpyxl convention).
    """
    merged = {}
    for rng in ws.merged_cells.ranges:
        master_val = ws.cell(rng.min_row, rng.min_col).value
        for r in range(rng.min_row, rng.max_row + 1):
            for c in range(rng.min_col, rng.max_col + 1):
                merged[(r, c)] = master_val
    return merged


def _cell_val(ws, row: int, col: int, merged_map: dict) -> Any:
    """Get cell value resolving merges. row/col are 1-indexed."""
    if (row, col) in merged_map:
        return merged_map[(row, col)]
    return ws.cell(row, col).value


def _detect_header_band(ws, merged_map: dict, max_scan: int = 25
                        ) -> tuple[int, int]:
    """
    Detect the header band: one or more consecutive rows forming the header.

    Strategy:
      1. For each row in the scan range, compute a "keyword hit score" by
         checking how many known COLUMN_MAP variations appear in that row.
      2. The row (or band of consecutive rows) with the highest combined score
         is the header.
      3. Handles multi-level headers by flattening stacked rows.

    Returns (first_header_row, last_header_row), both 1-indexed.
    """
    # Collect all known header keywords
    all_keywords = set()
    for variations in COLUMN_MAP.values():
        for v in variations:
            all_keywords.add(_normalise_header_text(v))

    # Score each row by keyword hits
    row_scores: dict[int, float] = {}
    row_cell_counts: dict[int, int] = {}
    max_col = min(ws.max_column or 1, 200)  # cap to avoid scanning thousands of cols

    for row_idx in range(1, min(max_scan + 1, (ws.max_row or 1) + 1)):
        score = 0.0
        cell_count = 0
        for col_idx in range(1, max_col + 1):
            val = _cell_val(ws, row_idx, col_idx, merged_map)
            if val is None:
                continue
            text = _normalise_header_text(str(val))
            if not text:
                continue
            cell_count += 1
            # Skip very short values for keyword matching — they produce
            # false positives (e.g. "Y" is a substring of "recovery")
            if len(text) < 3:
                continue
            for kw in all_keywords:
                # For containment, require the shorter string to be >= 4 chars
                shorter = min(len(text), len(kw))
                if shorter >= 4 and (kw in text or text in kw):
                    score += 1.0
                    break
                elif text == kw:
                    score += 1.0
                    break
                elif len(text) >= 4 and _fuzzy_score(text, kw) > 0.80:
                    score += 0.7
                    break
        row_scores[row_idx] = score
        row_cell_counts[row_idx] = cell_count

    if not row_scores:
        return (1, 1)

    # Find the best row by keyword score, breaking ties by cell count
    best_row = max(row_scores, key=lambda r: (row_scores[r], row_cell_counts.get(r, 0)))
    best_score = row_scores[best_row]

    if best_score == 0:
        # Fallback: use the row with the most non-empty cells
        best_row = max(row_cell_counts, key=lambda r: row_cell_counts[r])
        return (best_row, best_row)

    # Check if adjacent rows are part of a multi-level header band.
    # A row is part of the band only if it has SUBSTANTIAL keyword hits —
    # a single incidental match (e.g. data cell containing "supplier") is noise.
    # Require at least 2 keyword hits or 20% of the best row's score.
    header_start = best_row
    header_end = best_row
    min_adjacent_score = max(2.0, best_score * 0.15)

    # Check row above
    if best_row > 1:
        above = best_row - 1
        above_score = row_scores.get(above, 0)
        above_cells = row_cell_counts.get(above, 0)
        if above_score >= min_adjacent_score and above_cells >= 3:
            header_start = above

    # Check row below
    below = best_row + 1
    if below <= (ws.max_row or 1):
        below_score = row_scores.get(below, 0)
        below_cells = row_cell_counts.get(below, 0)
        if below_score >= min_adjacent_score and below_cells >= 3:
            header_end = below

    return (header_start, header_end)


def _flatten_header_band(ws, merged_map: dict,
                         start_row: int, end_row: int) -> list[str]:
    """
    Flatten a multi-row header band into a single list of header strings.
    For each column, concatenate non-empty values from all rows in the band,
    separated by ' — ', to produce a composite header.
    """
    max_col = min(ws.max_column or 1, 200)

    # Find the actual rightmost column that has content in the header band
    rightmost = 0
    for row_idx in range(start_row, end_row + 1):
        for col_idx in range(max_col, 0, -1):
            val = _cell_val(ws, row_idx, col_idx, merged_map)
            if val is not None and str(val).strip():
                rightmost = max(rightmost, col_idx)
                break
    if rightmost == 0:
        rightmost = max_col

    headers: list[str] = []
    for col_idx in range(1, rightmost + 1):
        parts = []
        for row_idx in range(start_row, end_row + 1):
            val = _cell_val(ws, row_idx, col_idx, merged_map)
            if val is not None:
                text = str(val).strip()
                if text and text not in parts:
                    parts.append(text)
        combined = " — ".join(parts) if parts else ""
        headers.append(combined)

    return headers


# ─── Column Matching ─────────────────────────────────────────────────

def _match_columns(headers: list[str]) -> tuple[dict[str, int], list[tuple[int, str]], list[str]]:
    """
    Match raw (possibly multi-level) headers to semantic field names.

    Uses a three-pass strategy:
      1. Exact match (normalised)
      2. Containment match (keyword in header or header in keyword)
      3. Fuzzy match (SequenceMatcher > 0.75 threshold)

    Returns:
        mapped:          {field_name: col_index}   (0-indexed)
        customer_cols:   [(col_index, header_text)]
        parse_warnings:  list[str]
    """
    mapped: dict[str, int] = {}
    matched_indices: set[int] = set()
    parse_warnings: list[str] = []

    headers_norm = [_normalise_header_text(h) for h in headers]

    # --- Pass 1: Exact normalised match ---
    for field, variations in COLUMN_MAP.items():
        if field in mapped:
            continue
        for var in variations:
            var_norm = _normalise_header_text(var)
            for idx, h_norm in enumerate(headers_norm):
                if idx in matched_indices or not h_norm:
                    continue
                if h_norm == var_norm:
                    mapped[field] = idx
                    matched_indices.add(idx)
                    break
            if field in mapped:
                break

    # --- Pass 2: Containment match (keyword in header) ---
    for field, variations in COLUMN_MAP.items():
        if field in mapped:
            continue
        best_idx = None
        best_specificity = 0  # prefer longest matching variation
        for var in variations:
            var_norm = _normalise_header_text(var)
            if len(var_norm) < 3:
                continue  # skip very short patterns to avoid false matches
            for idx, h_norm in enumerate(headers_norm):
                if idx in matched_indices or not h_norm:
                    continue
                if var_norm in h_norm or h_norm in var_norm:
                    specificity = len(var_norm)
                    if specificity > best_specificity:
                        best_specificity = specificity
                        best_idx = idx
            if best_idx is not None:
                break
        if best_idx is not None:
            mapped[field] = best_idx
            matched_indices.add(best_idx)

    # --- Pass 3: Fuzzy match for remaining unmapped fields ---
    FUZZY_THRESHOLD = 0.75
    for field, variations in COLUMN_MAP.items():
        if field in mapped:
            continue
        best_idx = None
        best_score = FUZZY_THRESHOLD
        for var in variations:
            var_norm = _normalise_header_text(var)
            for idx, h_norm in enumerate(headers_norm):
                if idx in matched_indices or not h_norm:
                    continue
                score = _fuzzy_score(h_norm, var_norm)
                if score > best_score:
                    best_score = score
                    best_idx = idx
        if best_idx is not None:
            mapped[field] = best_idx
            matched_indices.add(best_idx)
            parse_warnings.append(
                f"Fuzzy-matched column '{headers[best_idx]}' → field '{field}' "
                f"(score={best_score:.2f})"
            )

    # --- Disambiguation: "allocation" vs "allocation responsible" ---
    # If both fields mapped to the same column, break the tie by specificity.
    if ("allocation" in mapped and "allocation_responsible" in mapped
            and mapped["allocation"] == mapped["allocation_responsible"]):
        # Keep allocation_responsible (more specific), unmap allocation
        del mapped["allocation"]

    # --- Report missing critical fields ---
    critical_fields = ["product_family_desc", "plant_location",
                       "coverage_without_mitigation", "coverage_with_mitigation"]
    for f in critical_fields:
        if f not in mapped:
            parse_warnings.append(f"Missing expected column: {f}")

    # --- Classify unmatched columns ---
    customer_cols: list[tuple[int, str]] = []
    time_week_pattern = re.compile(
        r'^(cw\s*\d+|kw\s*\d+|w\s*\d+|\d{1,2}|q[1-4]|q\+?\d)$',
        re.IGNORECASE
    )

    for idx, h in enumerate(headers):
        if idx in matched_indices:
            continue
        h_str = str(h).strip()
        h_norm = _normalise_header_text(h_str)
        if not h_norm:
            continue
        if time_week_pattern.match(h_norm):
            mapped[f"_time_{h_str}"] = idx
            matched_indices.add(idx)
        else:
            customer_cols.append((idx, h_str))
            matched_indices.add(idx)

    return mapped, customer_cols, parse_warnings


# ─── Data Row Filtering ──────────────────────────────────────────────

def _is_data_row(cell_values: list[Any], mapped: dict[str, int]) -> bool:
    """
    Determine whether a row contains actual data vs. being a separator,
    subtotal, or decoration row.
    """
    # All empty
    non_empty = [v for v in cell_values if v is not None and str(v).strip()]
    if not non_empty:
        return False

    # Very few cells filled (likely a section header or separator)
    if len(non_empty) < 2:
        first_val = str(non_empty[0]).strip()
        if _SKIP_ROW_MARKERS.match(first_val):
            return False
        # Single cell could be a merged PG label — allow if PG field column is filled
        pg_idx = mapped.get("product_family_desc")
        if pg_idx is not None and pg_idx < len(cell_values):
            if cell_values[pg_idx] is not None and str(cell_values[pg_idx]).strip():
                return True
        return False

    # Check if any core field (plant, supplier, coverage) has a value
    core_fields = ["plant_location", "supplier_text", "coverage_with_mitigation",
                   "coverage_without_mitigation", "action_comment", "root_cause",
                   "customer_affected"]
    for field in core_fields:
        idx = mapped.get(field)
        if idx is not None and idx < len(cell_values):
            val = cell_values[idx]
            if val is not None and str(val).strip():
                return True

    # If we have at least 3 non-empty cells, accept it
    return len(non_empty) >= 3


# ─── Main Parser ─────────────────────────────────────────────────────

def parse_dashboard_update(filepath: str, sheet_name: str = "Dashboard_Update") -> dict[str, Any]:
    """
    Parse the Dashboard_Update worksheet from an Excel file.

    Handles:
      - Multi-row headers (auto-detected and flattened)
      - Headers not starting at row 1
      - Merged cells (both in headers and data)
      - Inconsistent formatting

    Each row includes:
      - coverage_cw_without_mitigation: (year, week) or None
      - coverage_cw_with_mitigation:    (year, week) or None
      - recovery_cw:                    (year, week) or None
    """
    path = Path(filepath)
    if not path.exists():
        raise FileNotFoundError(f"File not found: {filepath}")

    wb = openpyxl.load_workbook(str(path), read_only=False, data_only=True)
    parse_warnings: list[str] = []

    # ── Sheet selection (case-insensitive, partial match) ────────────
    ws = None
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
    else:
        for sn in wb.sheetnames:
            if sn.lower().strip() == sheet_name.lower().strip():
                ws = wb[sn]
                parse_warnings.append(f"Sheet name case mismatch: found '{sn}'")
                break
        if ws is None:
            # Try partial match
            for sn in wb.sheetnames:
                if sheet_name.lower() in sn.lower() or sn.lower() in sheet_name.lower():
                    ws = wb[sn]
                    parse_warnings.append(f"Sheet partial match: found '{sn}' for '{sheet_name}'")
                    break
        if ws is None:
            ws = wb[wb.sheetnames[0]]
            parse_warnings.append(
                f"Sheet '{sheet_name}' not found. Available: {wb.sheetnames}. "
                f"Using first sheet: '{wb.sheetnames[0]}'"
            )

    # ── Merged cell map ─────────────────────────────────────────────
    merged_map = _get_merged_cell_map(ws)

    # ── Header detection ────────────────────────────────────────────
    header_start, header_end = _detect_header_band(ws, merged_map)
    raw_headers = _flatten_header_band(ws, merged_map, header_start, header_end)

    if header_start != header_end:
        parse_warnings.append(
            f"Multi-row header detected: rows {header_start}–{header_end} "
            f"({header_end - header_start + 1} rows flattened)"
        )

    # ── Column matching ─────────────────────────────────────────────
    mapped, customer_cols, match_warnings = _match_columns(raw_headers)
    parse_warnings.extend(match_warnings)
    customer_col_names = [name for _, name in customer_cols]
    current_year = datetime.now().year

    # ── Parse data rows ─────────────────────────────────────────────
    rows: list[dict] = []
    data_start = header_end + 1
    n_cols = len(raw_headers)

    def _get_cell(row_idx: int, col_idx_0: int) -> Any:
        """Get cell value. col_idx_0 is 0-based."""
        col_1 = col_idx_0 + 1
        return _cell_val(ws, row_idx, col_1, merged_map)

    max_row = ws.max_row or 1

    for row_idx in range(data_start, max_row + 1):
        cell_values = [_get_cell(row_idx, c) for c in range(n_cols)]

        if not _is_data_row(cell_values, mapped):
            continue

        row_obj: dict[str, Any] = {}

        for field, col_idx in mapped.items():
            if field.startswith("_time_"):
                continue
            if col_idx < n_cols:
                raw = cell_values[col_idx]
                normaliser = _NORMALISERS.get(field, _norm_str)
                row_obj[field] = normaliser(raw)
            else:
                row_obj[field] = None

        # ── Parse coverage CW tuples ────────────────────────────────
        cov_wo = row_obj.get("coverage_without_mitigation")
        cov_w  = row_obj.get("coverage_with_mitigation")
        recovery = row_obj.get("recovery_week")

        row_obj["coverage_cw_without_mitigation"] = parse_cw(cov_wo, current_year)
        row_obj["coverage_cw_with_mitigation"] = parse_cw(cov_w, current_year)
        row_obj["recovery_cw"] = parse_cw(recovery, current_year)

        # ── Customer impact columns ─────────────────────────────────
        customer_impact: dict[str, Any] = {}
        for col_idx, cust_name in customer_cols:
            if col_idx < n_cols:
                customer_impact[cust_name] = _norm_str(cell_values[col_idx])
            else:
                customer_impact[cust_name] = None
        row_obj["customer_impact"] = customer_impact

        # ── Raw preservation ────────────────────────────────────────
        raw_columns: dict[str, Any] = {}
        for idx, header in enumerate(raw_headers):
            if header and idx < len(cell_values):
                raw_columns[header] = cell_values[idx]
        row_obj["raw_columns"] = raw_columns

        rows.append(row_obj)

    wb.close()

    product_groups = _group_by_product(rows)

    return {
        "rows": rows,
        "product_groups": product_groups,
        "customer_columns": customer_col_names,
        "warnings": parse_warnings,
        "metadata": {
            "filename": path.name,
            "sheet_name": ws.title,
            "header_rows": (header_start, header_end),
            "total_rows": len(rows),
            "total_columns": n_cols,
            "mapped_columns": len([k for k in mapped if not k.startswith("_time_")]),
            "customer_impact_columns": len(customer_col_names),
        },
    }


def _group_by_product(rows: list[dict]) -> list[dict]:
    groups: dict[str, dict] = {}
    order: list[str] = []
    for row in rows:
        desc = row.get("product_family_desc") or "Unknown"
        code = row.get("product_family_code") or ""
        key = f"{desc}|{code}"
        if key not in groups:
            groups[key] = {
                "product_family_desc": desc,
                "product_family_code": code,
                "rows": [],
            }
            order.append(key)
        groups[key]["rows"].append(row)
    return [groups[k] for k in order]


# ─── Convenience: Quick Summary for LLM ──────────────────────────────

def summarise_for_prompt(parsed: dict, max_chars: int = 12000) -> str:
    if not parsed.get("rows"):
        return ""

    current_cw = get_current_cw()
    lines = [
        "GLOBAL FULFILMENT DASHBOARD — Parsed Data Summary",
        f"Source: {parsed['metadata']['filename']}",
        f"Total risk rows: {parsed['metadata']['total_rows']}",
        f"Product groups: {len(parsed['product_groups'])}",
        f"Current CW: CW{current_cw[1]}/{current_cw[0]}",
        "",
    ]

    for pg in parsed["product_groups"]:
        desc = pg["product_family_desc"]
        code = pg["product_family_code"]
        header = f"{desc} ({code})" if code else desc
        lines.append(f"## {header} — {len(pg['rows'])} risk item(s)")

        for row in pg["rows"]:
            plant = row.get("plant_location") or "—"
            supplier = row.get("supplier_text") or "—"
            root = row.get("root_cause") or "—"
            cov_wo = row.get("coverage_without_mitigation") or "—"
            cov_w  = row.get("coverage_with_mitigation") or "—"
            comment = row.get("action_comment") or ""

            lines.append(f"  Plant: {plant} | Supplier: {supplier}")
            lines.append(f"  Coverage w/o mitigation: {cov_wo} | w/ mitigation: {cov_w}")
            lines.append(f"  Root cause: {root}")

            impacts = row.get("customer_impact", {})
            affected = [k for k, v in impacts.items() if v]
            if affected:
                lines.append(f"  Customers affected: {', '.join(affected[:10])}")
            if comment:
                lines.append(f"  Action/Comment: {comment[:200]}")
            lines.append("")

    text = "\n".join(lines)
    if len(text) > max_chars:
        text = text[:max_chars - 50] + "\n  ... [truncated for token budget]"
    return text
